from __future__ import annotations

import json
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from config import Settings
from paths import (
    CODE_EXTS,
    TEXT_EXTS,
    WHITELIST_EXTS,
    catalog_root,
    detect_branch,
    rel_path,
    should_skip_dir,
    should_skip_file,
)

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None  # type: ignore[misc, assignment]

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None  # type: ignore[misc, assignment]

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[misc, assignment]


@dataclass
class IndexEntry:
    path: str
    name: str
    ext: str | None
    size: int
    mtime: str
    branch: str | None
    text: str = ""


def _code_path_allowed(rel: str) -> bool:
    low = rel.lower().replace("\\", "/")
    return any(
        part in low
        for part in (
            "cursor/",
            "claude/",
            "aikivaviora/",
            "02modules/",
            "python_kash/",
        )
    )


class CatalogIndex:
    def __init__(self) -> None:
        self._entries: list[IndexEntry] = []
        self._lock = threading.Lock()

    @property
    def count(self) -> int:
        return len(self._entries)

    def clear(self) -> None:
        with self._lock:
            self._entries = []
        try:
            self._cache_path().unlink(missing_ok=True)
        except Exception:
            pass

    def entries(self) -> list[IndexEntry]:
        with self._lock:
            return list(self._entries)

    def _cache_path(self) -> Path:
        return Path(__file__).resolve().parent.parent / "05data" / "aviora_catalog_index.json"

    def save_cache(self, catalog_root: str) -> None:
        path = self._cache_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        with self._lock:
            payload = {
                "catalog_root": catalog_root,
                "entries": [
                    {
                        "path": e.path,
                        "name": e.name,
                        "ext": e.ext,
                        "size": e.size,
                        "mtime": e.mtime,
                        "branch": e.branch,
                        "text": e.text[:8000],
                    }
                    for e in self._entries
                ],
            }
        path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    def load_cache(self, catalog_root: str) -> int:
        path = self._cache_path()
        if not path.is_file():
            return 0
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return 0
        if (payload.get("catalog_root") or "").lower() != str(catalog_root).lower():
            return 0
        loaded: list[IndexEntry] = []
        for row in payload.get("entries") or []:
            loaded.append(
                IndexEntry(
                    path=row["path"],
                    name=row["name"],
                    ext=row.get("ext"),
                    size=int(row.get("size") or 0),
                    mtime=row.get("mtime") or "",
                    branch=row.get("branch"),
                    text=row.get("text") or "",
                )
            )
        with self._lock:
            self._entries = loaded
        return len(loaded)

    def _extract_pdf(self, path: Path, max_chars: int = 8000) -> str:
        if PdfReader is None:
            return ""
        try:
            reader = PdfReader(str(path))
            parts: list[str] = []
            for page in reader.pages[:30]:
                parts.append(page.extract_text() or "")
                if sum(len(p) for p in parts) > max_chars:
                    break
            return "\n".join(parts)[:max_chars]
        except Exception:
            return ""

    def _extract_docx(self, path: Path, max_chars: int = 8000) -> str:
        if DocxDocument is None:
            return ""
        try:
            doc = DocxDocument(str(path))
            text = "\n".join(p.text for p in doc.paragraphs if p.text)
            return text[:max_chars]
        except Exception:
            return ""

    def _extract_xlsx(self, path: Path, max_chars: int = 4000) -> str:
        if load_workbook is None:
            return ""
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.worksheets[:3]:
                for row in sheet.iter_rows(max_row=50, values_only=True):
                    parts.append(
                        " ".join(str(c) for c in row if c is not None)
                    )
            wb.close()
            return "\n".join(parts)[:max_chars]
        except Exception:
            return ""

    def _read_text_file(self, path: Path, max_bytes: int) -> str:
        for enc in ("utf-8", "cp1251"):
            try:
                data = path.read_bytes()[:max_bytes]
                return data.decode(enc, errors="replace")
            except Exception:
                continue
        return ""

    def collect_files(self, settings: Settings) -> list[Path]:
        root = catalog_root(settings)
        files: list[Path] = []
        for dirpath, dirnames, filenames in root.walk(top_down=True):
            dirnames[:] = [
                d
                for d in dirnames
                if not should_skip_dir(d, settings)
            ]
            for name in filenames:
                if should_skip_file(name, settings):
                    continue
                p = Path(dirpath) / name
                ext = p.suffix.lower()
                rel = rel_path(root, p)
                if ext in CODE_EXTS:
                    if not _code_path_allowed(rel):
                        continue
                if ext in WHITELIST_EXTS or ext in TEXT_EXTS or ext in CODE_EXTS:
                    files.append(p)
        return files

    def build(
        self,
        settings: Settings,
        *,
        cancel_check,
        on_progress,
    ) -> dict[str, Any]:
        from datetime import datetime, timezone

        root = catalog_root(settings)
        files = self.collect_files(settings)
        total = len(files)
        new_entries: list[IndexEntry] = []
        max_read = settings.max_file_read_mb * 1024 * 1024

        for i, path in enumerate(files, start=1):
            if cancel_check():
                return {"cancelled": True, "indexed": len(new_entries), "total": total}
            rel = rel_path(root, path)
            on_progress(i, total, rel)
            ext = path.suffix.lower()
            st = path.stat()
            mtime = datetime.fromtimestamp(
                st.st_mtime, tz=timezone.utc
            ).isoformat()
            text = ""
            if ext in CODE_EXTS:
                text = rel
            elif ext in TEXT_EXTS:
                text = self._read_text_file(path, max_read)
            elif ext == ".pdf":
                text = self._extract_pdf(path)
            elif ext == ".docx":
                text = self._extract_docx(path)
            elif ext == ".xlsx":
                text = self._extract_xlsx(path)
            new_entries.append(
                IndexEntry(
                    path=rel,
                    name=path.name,
                    ext=ext or None,
                    size=st.st_size,
                    mtime=mtime,
                    branch=detect_branch(rel),
                    text=text,
                )
            )

        with self._lock:
            self._entries = new_entries
        try:
            self.save_cache(str(root))
        except Exception:
            pass
        return {"cancelled": False, "indexed": len(new_entries), "total": total}

    def search(
        self,
        q: str,
        *,
        ext: str | None = None,
        branch: str | None = None,
        limit: int = 100,
    ) -> list[dict[str, Any]]:
        q_low = q.strip().lower()
        if not q_low:
            return []
        results: list[dict[str, Any]] = []
        for e in self.entries():
            if ext and (e.ext or "") != ext.lower():
                continue
            if branch and (e.branch or "") != branch:
                continue
            hay = f"{e.path} {e.name} {e.text}".lower()
            if q_low and q_low not in hay:
                continue
            snippet = ""
            if q_low and e.text:
                idx = e.text.lower().find(q_low)
                if idx >= 0:
                    start = max(0, idx - 40)
                    snippet = e.text[start : start + 120]
            results.append(
                {
                    "path": e.path,
                    "name": e.name,
                    "ext": e.ext,
                    "size": e.size,
                    "mtime": e.mtime,
                    "branch": e.branch,
                    "snippet": snippet,
                }
            )
            if len(results) >= limit:
                break
        return results


INDEX = CatalogIndex()
